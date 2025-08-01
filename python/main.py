from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from collections import defaultdict
import mysql.connector
import logging
import os
from datetime import datetime
from config import DB_CONFIG

app = FastAPI()

# Add CORS for frontend access
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve output files statically at /download
app.mount("/download", StaticFiles(directory="output"), name="download")

# Setup logging
logging.basicConfig(level=logging.INFO)

def safe_set_cell(ws, row, col, value):
    """Safely set cell value, avoiding merged cells"""
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value
    else:
        logging.warning(f"Skipped merged cell at ({row}, {col})")
@app.get("/download-excel")
def download_excel():
    """Generate and directly return Excel file for download"""
    try:
        # Connect to database
        logging.info("Connecting to database...")
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)

        # Fetch data
        cursor.execute("""
            SELECT year, programme_name, programme_code, no_of_seats, no_of_students
            FROM response_2_1_1
            WHERE criteria_code = '020201020101'
            ORDER BY year ASC, programme_name ASC
        """)
        rows = cursor.fetchall()

        if not rows:
            raise HTTPException(status_code=404, detail="No data found")

        # Group by year
        data_by_year = defaultdict(list)
        for row in rows:
            data_by_year[str(row['year'])].append(row)

        # Load template
        template_path = 'templates/criteria211Template.xlsx'
        if not os.path.exists(template_path):
            raise HTTPException(status_code=500, detail="Template file not found")
        
        wb = load_workbook(template_path)
        ws = wb.active

        # Year mapping
        available_years = sorted([year for year in data_by_year.keys()], reverse=True)[:5]
        year_sections = {
            'Year - 1': {'start_row': 4, 'years': [available_years[0]] if len(available_years) > 0 else []},
            'Year - 2': {'start_row': 12, 'years': [available_years[1]] if len(available_years) > 1 else []},
            'Year - 3': {'start_row': 20, 'years': [available_years[2]] if len(available_years) > 2 else []},
            'Year - 4': {'start_row': 27, 'years': [available_years[3]] if len(available_years) > 3 else []},
            'Year - 5': {'start_row': 37, 'years': [available_years[4]] if len(available_years) > 4 else []},
        }

        total_records = 0
        for section_name, section_info in year_sections.items():
            start_row = section_info['start_row']
            section_years = section_info['years']
            
            for year in section_years:
                year_data = data_by_year.get(year, [])
                for i, record in enumerate(year_data):
                    row_num = start_row + i
                    safe_set_cell(ws, row_num, 1, record['programme_name'])
                    safe_set_cell(ws, row_num, 2, record['programme_code'])
                    safe_set_cell(ws, row_num, 3, record['no_of_seats'])
                    safe_set_cell(ws, row_num, 4, record['no_of_students'])
                    total_records += 1

        # Save file
        os.makedirs('output', exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"criteria_2_1_1_enrolment_{timestamp}.xlsx"
        output_path = os.path.join('output', filename)
        wb.save(output_path)

        logging.info(f"Generated Excel: {output_path} with {total_records} rows")

        # ✅ Immediately return the file for download
        return FileResponse(
            path=output_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename
        )

    except mysql.connector.Error as e:
        logging.error(f"Database error: {e}")
        raise HTTPException(status_code=500, detail="Database connection failed")
    except Exception as e:
        logging.error(f"Unexpected error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
